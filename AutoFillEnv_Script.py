import fitz
import openpyxl
import os
import traceback
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

def run_process():
    try:
        # 1. Capture Inputs
        excel_path = excel_entry.get()
        template_path = pdf_entry.get()
        output_dir = folder_entry.get()
        font_filename = "MyriadPro-Semibold.otf"
        
        # Validation
        if not all([excel_path, template_path, output_dir]):
            messagebox.showerror("Error", "All paths must be selected.")
            return

        if not os.path.exists(font_filename):
            messagebox.showerror("Error", f"Font file '{font_filename}' not found in the same folder as this script.")
            return

        # 2. Extract Template Info
        src = fitz.open(template_path)
        first_page = src[0] # Explicitly get first page
        first_annot = next(first_page.annots(), None)
        
        if not first_annot:
            messagebox.showerror("Error", "No comment box found on Page 1 of PDF.")
            src.close()
            return
        
        rect_coords = [first_annot.rect.x0, first_annot.rect.y0, first_annot.rect.x1, first_annot.rect.y1]
        page_rect = first_page.rect
        src.close()

        # 3. Setup Processing
        output_doc = fitz.open()
        final_rect = fitz.Rect(rect_coords)
        template_doc = fitz.open(template_path)
        
        with open(font_filename, "rb") as f:
            font_bytes = f.read()

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        process_all = range_var.get()
        try:
            start_row = int(start_entry.get() or 2)
            end_row = int(end_entry.get()) if not process_all and end_entry.get() else ws.max_row
        except:
            start_row, end_row = 2, ws.max_row
        
        rows = list(ws.iter_rows(min_row=start_row, max_row=end_row, values_only=True))
        total = len(rows)
        progress_bar['maximum'] = total
        
        # 4. Main Loop
        for i, row in enumerate(rows, 1):
            def get_val(idx): 
                try:
                    val = row[idx]
                    return str(val) if val is not None else ""
                except IndexError:
                    return ""
            
            c2, e2, f2, g2, h2 = get_val(2), get_val(4), get_val(5), get_val(6), get_val(7)
            text_content = f"Human Resources\n{c2}\n{e2}\n{f2}, {g2} {h2}"

            new_page = output_doc.new_page(width=page_rect.width, height=page_rect.height)
            new_page.show_pdf_page(new_page.rect, template_doc, 0)
            
            new_page.insert_font(fontname="my_font", fontbuffer=font_bytes)
            new_page.insert_textbox(
                final_rect, 
                text_content, 
                fontsize=14, 
                fontname="my_font", 
                lineheight=1.5,
                align=0
            )

            # Update UI
            progress_bar['value'] = i
            status_label.config(text=f"Processed {i} of {total}")
            root.update()

        # 5. Finalize
        output_path = os.path.join(output_dir, "1_Envelope_Filled.pdf")
        output_doc.save(output_path)
        template_doc.close()
        output_doc.close()
        
        messagebox.showinfo("Success", "Process Complete!")

    except Exception:
        # 🛑 THIS WILL POP UP A WINDOW WITH THE REAL ERROR
        error_msg = traceback.format_exc()
        messagebox.showerror("Fatal Error Traceback", error_msg)

# --- UI Setup ---
root = tk.Tk()
root.title("Envelope PDF Tool")
root.geometry("820x250")

# Standard Browse Helpers
def set_path(entry):
    path = filedialog.askopenfilename() if entry != folder_entry else filedialog.askdirectory()
    if path: 
        entry.delete(0, tk.END)
        entry.insert(0, path)

# Toggle logic for the checkboxes
def toggle_range_mode():
    if range_var.get():
        custom_var.set(False)
        start_entry.config(state="disabled")
        end_entry.config(state="disabled")
    else:
        custom_var.set(True)
        start_entry.config(state="normal")
        end_entry.config(state="normal")

def toggle_custom_mode():
    if custom_var.get():
        range_var.set(False)
        start_entry.config(state="normal")
        end_entry.config(state="normal")
    else:
        range_var.set(True)
        start_entry.config(state="disabled")
        end_entry.config(state="disabled")

# Grid Container for Inputs
input_frame = tk.Frame(root)
input_frame.pack(pady=15, padx=15, fill="x")

# Row 1: Source Excel
tk.Label(input_frame, text="Source Excel:", anchor="w").grid(row=0, column=0, sticky="w", pady=5)
excel_entry = tk.Entry(input_frame, width=75)
excel_entry.grid(row=0, column=1, padx=5, pady=5)
tk.Button(input_frame, text="Browse", command=lambda: set_path(excel_entry)).grid(row=0, column=2, padx=5, pady=5)

# Row 2: PDF Template
tk.Label(input_frame, text="PDF Template:", anchor="w").grid(row=1, column=0, sticky="w", pady=5)
pdf_entry = tk.Entry(input_frame, width=75)
pdf_entry.grid(row=1, column=1, padx=5, pady=5)
tk.Button(input_frame, text="Browse", command=lambda: set_path(pdf_entry)).grid(row=1, column=2, padx=5, pady=5)

# Row 3: Output Directory
tk.Label(input_frame, text="Output Directory:", anchor="w").grid(row=2, column=0, sticky="w", pady=5)
folder_entry = tk.Entry(input_frame, width=75)
folder_entry.grid(row=2, column=1, padx=5, pady=5)
tk.Button(input_frame, text="Browse", command=lambda: set_path(folder_entry)).grid(row=2, column=2, padx=5, pady=5)

# Bottom Options and Action Row
control_frame = tk.Frame(root)
control_frame.pack(fill="x", padx=15, pady=5)

# Left Side of Controls: Two Checkboxes and Ranges
range_frame = tk.Frame(control_frame)
range_frame.pack(side="left", anchor="w")

range_var = tk.BooleanVar(value=True)
custom_var = tk.BooleanVar(value=False)

# Checkbox 1: All Rows
cb_all = tk.Checkbutton(range_frame, text="All Rows", variable=range_var, command=toggle_range_mode)
cb_all.grid(row=0, column=0, padx=(0, 10))

# Checkbox 2: Custom Range
cb_custom = tk.Checkbutton(range_frame, text="Custom Range", variable=custom_var, command=toggle_custom_mode)
cb_custom.grid(row=0, column=1, padx=(0, 15))

# Range Entry Fields
tk.Label(range_frame, text="Start:").grid(row=0, column=2)
start_entry = tk.Entry(range_frame, width=5, state="disabled")
start_entry.insert(0, "2")
start_entry.grid(row=0, column=3, padx=5)

tk.Label(range_frame, text="End:").grid(row=0, column=4, padx=(10, 0))
end_entry = tk.Entry(range_frame, width=5, state="disabled")
end_entry.grid(row=0, column=5, padx=5)

# Right Side of Controls: Execution Button
tk.Button(control_frame, text="START", bg="green", fg="white", font=("Arial", 11, "bold"), width=12, command=run_process).pack(side="right", padx=5)

# Status Footer
footer_frame = tk.Frame(root)
footer_frame.pack(fill="x", padx=15, pady=(10, 0))

status_label = tk.Label(footer_frame, text="Ready", fg="blue", anchor="w")
status_label.pack(side="left")

progress_bar = ttk.Progressbar(footer_frame, length=450)
progress_bar.pack(side="right")

root.mainloop()
