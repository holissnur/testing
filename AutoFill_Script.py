import os
import sys
import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook
import fitz  # PyMuPDF

# --- CONFIGURATION ---
EXCEL_FILE = "data_sample_template.xlsx"
PDF_TEMPLATE = "sample_template.pdf"
OUTPUT_FOLDER = "Output_Sample"

# Create output folder if it doesn't exist
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# --- COLUMN TO PDF FIELD MAPPING ---
COLUMN_MAPPING = {
    "B": "Address",
    "C": "State",
    "D": "Type Of Registration"
}

def fill_pdf_form(template_path, output_path, data_dict):
    """Opens a PDF template, fills the fields based on data_dict, and saves it."""
    doc = fitz.open(template_path)
    
    for page in doc:
        for field in page.widgets():
            field_name = field.field_name
            
            if field_name in data_dict:
                val = data_dict[field_name]
                if val is None or val == "":
                    continue
                
                val_str = str(val).strip()
                
                if field.field_type == fitz.PDF_WIDGET_TYPE_CHECKBOX:
                    if val_str.lower() in ["true", "yes", "1", "checked", "on"]:
                        field.field_value = "Yes"
                    elif val_str.lower() in ["false", "no", "0", "unchecked", "off"]:
                        field.field_value = "Off"
                    else:
                        field.field_value = val_str
                        
                elif field.field_type == 5:
                    raw_export = field.on_state()
                    decoded_export = raw_export.replace("#20", " ")
                    if val_str == decoded_export or val_str == raw_export:
                        field.field_value = True
                    else:
                        field.field_value = False
                    
                else:
                    field.field_value = val_str
                
                field.update()
                
    doc.save(output_path)
    doc.close()

def process_data(root, progress_bar, status_label, percentage_label, close_button):
    """Reads Excel and fills PDFs while updating the GUI elements dynamically."""
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.active  
    
    start_row = 2
    end_row = ws.max_row
    total_rows = max(0, end_row - start_row + 1)
    
    if total_rows == 0:
        status_label.config(text="No rows found to process.")
        close_button.pack(pady=10) # Show button anyway to allow closing
        return

    progress_bar["maximum"] = total_rows

    for current_count, row_idx in enumerate(range(start_row, end_row + 1), start=1):
        row_data = {}
        address_value_for_name = "filled"
        
        for col_letter, pdf_field_name in COLUMN_MAPPING.items():
            cell_value = ws[f"{col_letter}{row_idx}"].value
            
            if cell_value is not None:
                row_data[pdf_field_name] = cell_value
                
                if col_letter == "B":
                    address_value_for_name = "".join(c for c in str(cell_value) if c.isalnum() or c in (' ', '_')).strip()
                    address_value_for_name = address_value_for_name.replace(" ", "_")

        filename = f"{row_idx - 1}_{address_value_for_name}_filled.pdf"

        if not row_data:
            progress_bar["value"] = current_count
            root.update()
            continue
            
        status_label.config(text=f"Generating: {filename}")
        percent = int((current_count / total_rows) * 100)
        percentage_label.config(text=f"{percent}% ({current_count}/{total_rows})")
        
        output_pdf_path = os.path.join(OUTPUT_FOLDER, filename)
        fill_pdf_form(PDF_TEMPLATE, output_pdf_path, row_data)
        
        progress_bar["value"] = current_count
        root.update()
        
    # Update UI to finished state
    status_label.config(text="Processing complete! 'Output_Sample' folder updated.")
    percentage_label.config(text="100% Completed")
    
    # Make the close button visible to the user
    close_button.pack(pady=(0, 10))

def main():
    root = tk.Tk()
    root.title("Process Status")
    
    # Increased vertical size slightly to make room for the emerging button
    root.geometry("450x180")
    root.resizable(False, False)
    
    status_label = ttk.Label(root, text="Initializing Excel data...", wraplength=420)
    status_label.pack(pady=(20, 5), padx=20, anchor="w")
    
    progress_bar = ttk.Progressbar(root, orient="horizontal", length=400, mode="determinate")
    progress_bar.pack(pady=5, padx=20)
    
    percentage_label = ttk.Label(root, text="0%")
    percentage_label.pack(pady=(0, 5), padx=20, anchor="e")
    
    # Pre-create the close button widget linked to root.destroy, but do NOT pack it yet
    close_button = ttk.Button(root, text="OK", command=root.destroy)
    
    # Pass the button object down to trigger layout placement inside the process loop
    root.after(100, lambda: process_data(root, progress_bar, status_label, percentage_label, close_button))
    
    root.mainloop()

if __name__ == "__main__":
    main()
